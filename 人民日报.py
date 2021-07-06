import requests
from bs4 import BeautifulSoup
import datetime
from openpyxl import load_workbook, Workbook
from os import path
import re


def save_html(file, news_rsp):
    with open(file, 'w', encoding="utf-8") as f:
        html_doc = str(news_rsp.content, 'utf-8')
        html_doc = html_doc.replace('http://paper.people.com.cn/rmrb/html/', 'http://localhost/rmrb/html/')
        f.write(html_doc)


def rmrb(start_dt=datetime.datetime.now().strftime('%Y-%m/%d'), end_dt=datetime.datetime.now().strftime('%Y-%m/%d'), key_words=''):
    '''
    搜索日期从 start_dt 到 end_dt 的新闻标题中有 key_words 的新闻
    :param start_dt:
    :param end_dt:
    :param key_words:
    :return:
    '''
    base_url = 'http://paper.people.com.cn/rmrb/html/'
    # base_url = 'http://localhost/rmrb/html/'
    news_dt = start_dt
    today = datetime.datetime.now().strftime('%Y%m%d')
    file = '新闻{0}.xlsx'.format(today)
    if path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    while news_dt <= end_dt:
        for idx in range(1, 21):
            try:
                # 20 个板块链接
                url_panel = f'{base_url}{news_dt}/nbs.D110000renmrb_{str(idx).zfill(2)}.htm'
                # print(url_panel)
                rsp = requests.get(url_panel)
                soup = BeautifulSoup(rsp.content, 'html.parser')
                a_tags = soup.find_all('a', href=re.compile('nw.'))
                for item in a_tags:
                    news_url = base_url + news_dt + '/' + item.attrs.get('href')
                    new_title = item.contents[0]
                    news_rsp = requests.get(news_url)
                    # 保存网页
                    # save_html(item.attrs['href'], news_rsp)

                    news_soup = BeautifulSoup(news_rsp.content, 'html.parser')
                    news_artical = news_soup.find('div', attrs={"class": "article"})
                    news_content = news_artical.find_all('p', text=re.compile(key_words), limit=1)
                    if len(news_content) > 0:
                        ws.append((news_url, new_title))
            except Exception as e:
                print(e)
        d = datetime.datetime.strptime(news_dt, '%Y-%m/%d')
        news_dt = (d + datetime.timedelta(days=1)).strftime('%Y-%m/%d')
        wb.save(file)
        wb.close()


if __name__ == '__main__':
    start_dt = '2021-01/26'  # 2021-01/19
    end_dt = '2021-01/26'
    key_words = '课题'
    rmrb(start_dt, end_dt, key_words)
