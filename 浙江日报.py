import requests
from bs4 import BeautifulSoup
import datetime
from openpyxl import load_workbook, Workbook
from os import path


def zjrb(start_dt=datetime.datetime.now().strftime('%Y-%m/%d'), end_dt=datetime.datetime.now().strftime('%Y-%m/%d'), key_words=''):
    base_url = 'http://zjrb.zjol.com.cn/html/'
    news_dt = start_dt
    today = datetime.datetime.now().strftime('%Y%m%d')
    file = '新闻{0}.xlsx'.format(today)
    if path.exists(file):
        wb = load_workbook(file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
    while news_dt <= end_dt:
        for idx in range(3, 19):
            # 3 - 18 个板块链接
            url_panel = '{2}{1}/node_{0}.htm'.format(str(idx), news_dt, base_url)
            r = requests.get(url_panel)
            soup = BeautifulSoup(r.content, 'html.parser')
            a_tags = soup.find_all('a')
            for item in a_tags:
                try:
                    if item.attrs.get('href')[:7] == 'content':
                        news_title = item.contents[0]
                        news_url = '{2}{1}/{0}'.format(item.attrs.get('href'), news_dt, base_url)
                        news_rsp = requests.get(news_url)
                        news_soup = BeautifulSoup(news_rsp.content, 'html.parser')
                        news_content = str(news_soup.find(name='founder-content'))
                        if news_content.__contains__(key_words):
                        # if item.attrs.get('href')[:7] == 'content' and item.contents[0].__contains__(key_words):
                            ws.append((news_url, news_title))
                except Exception:
                    pass
        d = datetime.datetime.strptime(news_dt, '%Y-%m/%d')
        news_dt = (d + datetime.timedelta(days=1)).strftime('%Y-%m/%d')
    wb.save(file)
    wb.close()


if __name__ == '__main__':
    start_dt = '2021-01/18'  # 2021-01/19
    end_dt = '2021-01/19'
    key_words = '的'
    zjrb(start_dt, end_dt, key_words)
